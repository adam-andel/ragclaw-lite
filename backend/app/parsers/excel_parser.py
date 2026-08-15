"""Excel document parser using openpyxl (read-only mode for memory efficiency)."""

from pathlib import Path

from openpyxl import load_workbook

from app.parsers.base import BaseParser, ParsedDocument, ParsedSection, ParserPluginMeta


class ExcelParser(BaseParser):
    """Parse .xlsx files, producing one section per worksheet.

    Uses read_only=True to stream rows without loading the whole workbook
    into memory — important for large spreadsheets.
    """

    _MAX_ROWS_PER_SHEET = 1000

    def extensions(self) -> list[str]:
        return ["xlsx", "xls"]

    @classmethod
    def plugin_meta(cls) -> ParserPluginMeta:
        return ParserPluginMeta(
            name="excel",
            display_name="parser.excel.name",
            description="parser.excel.desc",
            category="office",
            extensions=["xlsx", "xls"],
        )

    def parse(self, file_path: Path) -> ParsedDocument:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        sections: list[ParsedSection] = []
        sheet_count = 0

        for sheet_idx, ws in enumerate(wb.worksheets):
            sheet_count += 1
            row_count = 0
            row_lines: list[str] = []

            for row in ws.iter_rows(values_only=True):
                cells = [self._stringify(c) for c in row]
                if any(cells):
                    row_lines.append("\t".join(cells))
                    row_count += 1
                if row_count >= self._MAX_ROWS_PER_SHEET:
                    break

            if row_lines:
                sections.append(ParsedSection(
                    level=1,
                    heading=ws.title or f"Sheet{sheet_idx + 1}",
                    content="\n".join(row_lines),
                    page=sheet_idx + 1,
                ))

        wb.close()

        if not sections:
            sections.append(ParsedSection(
                level=0, heading=file_path.stem, content="",
            ))

        return ParsedDocument(
            title=file_path.stem, file_type="xlsx", sections=sections,
            metadata={"sheet_count": sheet_count},
        )

    @staticmethod
    def _stringify(cell) -> str:
        if cell is None:
            return ""
        return str(cell)
